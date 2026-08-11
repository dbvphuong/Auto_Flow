from PyQt6.QtWidgets import (QWidget, QHBoxLayout, QVBoxLayout, QComboBox, 
                             QLabel, QSpinBox, QTextEdit, QPushButton, 
                             QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox, QSplitter,
                             QCheckBox, QLineEdit, QFileDialog, QGridLayout, QMessageBox, QTabWidget)
from PyQt6.QtCore import Qt, QThread, QTimer, QUrl
from PyQt6.QtGui import QDesktopServices, QIcon, QPixmap
import time
import random
import logging
import os

from data.database import SessionLocal
from data.models import Task, Account, VideoSession
from core.workers import AutomationWorker


from PyQt6.QtWidgets import QDialog, QDialogButtonBox

class PromptEditDialog(QDialog):
    def __init__(self, parent=None, session_name="", initial_text=""):
        super().__init__(parent)
        self.imported_prompt_files = []
        self.setWindowTitle(f"Nhập Prompt - {session_name}")
        self.resize(500, 400)
        
        layout = QVBoxLayout(self)
        
        lbl = QLabel(f"Nhập danh sách prompt cho {session_name} (mỗi dòng 1 prompt):")
        layout.addWidget(lbl)
        
        self.txt_edit = QTextEdit()
        self.txt_edit.setPlainText(initial_text)
        self.txt_edit.setPlaceholderText("Nhập prompt ở đây...")
        layout.addWidget(self.txt_edit)
        
        # Nút Import File
        self.btn_import = QPushButton("Import nhiều File (.txt, .xlsx)")
        self.btn_import.setToolTip("Mỗi file sẽ tạo một phiên riêng và tự đặt thư mục lưu")
        self.btn_import.clicked.connect(self.import_from_file)
        self.btn_import.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6; 
                color: white; 
                padding: 6px; 
                border-radius: 4px;
                border: none;
            }
            QPushButton:hover {
                background-color: #60a5fa;
            }
        """)
        layout.addWidget(self.btn_import)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
    def get_text(self):
        return self.txt_edit.toPlainText().strip()

    def get_imported_prompt_files(self):
        return self.imported_prompt_files
        
    def import_from_file(self):
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, "Chọn tệp prompt", "", "Text/Excel Files (*.txt *.xlsx *.xls *.csv)"
        )
        if not file_paths:
            return

        imported_files = []
        errors = []
        try:
            for file_path in file_paths:
                prompts = []
                try:
                    file_ext = os.path.splitext(file_path)[1].lower()
                    if file_ext == '.txt':
                        with open(file_path, 'r', encoding='utf-8-sig') as f:
                            prompts = [line.strip() for line in f if line.strip()]
                    elif file_ext == '.csv':
                        import csv
                        with open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
                            reader = csv.reader(f)
                            prompts = [row[0].strip() for row in reader if row and row[0].strip()]
                    elif file_ext in ('.xlsx', '.xls'):
                        from openpyxl import load_workbook
                        wb = load_workbook(file_path, read_only=True, data_only=True)
                        try:
                            sheet = wb.active
                            prompts = [str(row[0]).strip() for row in sheet.iter_rows(values_only=True)
                                       if row and row[0] is not None and str(row[0]).strip()]
                        finally:
                            wb.close()

                    if prompts:
                        imported_files.append((file_path, "\n".join(prompts)))
                    else:
                        errors.append(os.path.basename(file_path))
                except Exception as file_error:
                    logging.warning(f"[UI] Không đọc được file prompt {file_path}: {file_error}")
                    errors.append(os.path.basename(file_path))

            if not imported_files:
                QMessageBox.warning(self, "Cảnh báo", "Không tìm thấy prompt hợp lệ trong tệp.")
                return

            self.imported_prompt_files = imported_files
            if errors:
                QMessageBox.warning(
                    self,
                    "Cảnh báo",
                    "Không đọc được hoặc không có prompt hợp lệ trong: " + ", ".join(errors)
                )
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi đọc tệp: {str(e)}")

class FlowVideoView(QWidget):
    def __init__(self):
        super().__init__()
        self.thumbnail_cache = {}
        self.loading_thumbnails = {}
        self.page_size = 100
        self.current_page = 1
        self.total_pages = 1
        self.init_ui()

    def init_ui(self):
        main_layout = QHBoxLayout(self)
        
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(self.splitter)
        
        # Left Panel: Configuration
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 10, 0)
        
        # Banner tiêu đề "ẢNH" màu Gradient Xanh-Tím cực đẹp, nổi bật
        self.lbl_banner = QLabel("✨ ẢNH ✨")
        self.lbl_banner.setStyleSheet("""
            QLabel {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #1d4ed8, stop:0.5 #3b82f6, stop:1 #7c3aed);
                color: #ffffff;
                font-size: 22px;
                font-weight: 900;
                padding: 10px;
                border-radius: 8px;
                border: 2px solid rgba(255, 255, 255, 0.2);
                qproperty-alignment: 'AlignCenter';
                letter-spacing: 8px;
                margin-bottom: 10px;
            }
        """)
        left_layout.addWidget(self.lbl_banner)
        
        group_config = QGroupBox("Cấu hình & Prompts")
        config_layout = QGridLayout()
        config_layout.setSpacing(10)
        
        # Row 0
        config_layout.addWidget(QLabel("Model:"), 0, 0)
        config_layout.addWidget(QLabel("Chất lượng:"), 0, 1)
        
        # Row 1
        self.combo_model = QComboBox()
        self.combo_model.addItems(["Nano Banana Pro", "Nano Banana 2", "Imagen 4 (Leaving 6/16)"])
        config_layout.addWidget(self.combo_model, 1, 0)
        
        quality_layout = QHBoxLayout()
        self.chk_720p = QCheckBox("720p")
        self.chk_1080p = QCheckBox("1080p")
        self.chk_4k = QCheckBox("4K")
        self.chk_720p.setChecked(True)
        quality_layout.addWidget(self.chk_720p)
        quality_layout.addWidget(self.chk_1080p)
        quality_layout.addWidget(self.chk_4k)
        config_layout.addLayout(quality_layout, 1, 1)
        
        # Row 2
        config_layout.addWidget(QLabel("Tỷ lệ ảnh:"), 2, 0)
        config_layout.addWidget(QLabel("Số lượng video / prompt:"), 2, 1)
        
        # Row 3
        self.combo_ratio = QComboBox()
        self.combo_ratio.addItems(["16:9 Ngang", "9:16 Dọc", "4:3 Ngang", "3:4 Dọc", "1:1 Vuông"])
        config_layout.addWidget(self.combo_ratio, 3, 0)
        
        self.spin_images_per_prompt = QSpinBox()
        self.spin_images_per_prompt.setRange(1, 100)
        self.spin_images_per_prompt.setValue(1)
        config_layout.addWidget(self.spin_images_per_prompt, 3, 1)
        
        # Row 4
        config_layout.addWidget(QLabel("Số luồng chạy đồng thời:"), 4, 0)
        config_layout.addWidget(QLabel("Độ trễ giữa các luồng (s):"), 4, 1)
        
        # Row 5
        self.spin_threads = QSpinBox()
        self.spin_threads.setRange(1, 10)
        self.spin_threads.setValue(1)
        config_layout.addWidget(self.spin_threads, 5, 0)
        
        delay_layout = QHBoxLayout()
        self.spin_delay_min = QSpinBox()
        self.spin_delay_min.setRange(0, 3600)
        self.spin_delay_min.setValue(10)
        self.spin_delay_max = QSpinBox()
        self.spin_delay_max.setRange(0, 3600)
        self.spin_delay_max.setValue(20)
        delay_layout.addWidget(self.spin_delay_min)
        delay_layout.addWidget(QLabel("-"))
        delay_layout.addWidget(self.spin_delay_max)
        config_layout.addLayout(delay_layout, 5, 1)
        
        # Row 6
        config_layout.addWidget(QLabel("Chế độ tham chiếu:"), 6, 0)
        config_layout.addWidget(QLabel("Seed:"), 6, 1)
        
        # Row 7
        self.combo_ref_mode = QComboBox()
        self.combo_ref_mode.addItems(["Mặc định", "1 cho tất cả"])
        config_layout.addWidget(self.combo_ref_mode, 7, 0)
        
        seed_layout = QHBoxLayout()
        self.line_seed = QLineEdit()
        self.line_seed.setPlaceholderText("60475?")
        self.chk_lock_seed = QCheckBox("Khóa seed")
        seed_layout.addWidget(self.line_seed)
        seed_layout.addWidget(self.chk_lock_seed)
        config_layout.addLayout(seed_layout, 7, 1)
        
        group_config.setLayout(config_layout)
        left_layout.addWidget(group_config)
        
        # ----------------------------------------------------
        # QTabWidget bọc cấu hình nạp đơn & nạp phiên
        # ----------------------------------------------------
        # Chạy theo phiên (Batch) - Removed Tab Container
        batch_layout = QVBoxLayout()
        batch_layout.setContentsMargins(5, 5, 5, 5)
        batch_layout.setSpacing(8)
        
        self.table_sessions = QTableWidget(0, 7)
        self.table_sessions.setHorizontalHeaderLabels([" ", "Tên phiên", "Prompts", "Ảnh tham chiếu", "Thư mục lưu", "Tiến độ", "🔄"])
        self.table_sessions.setColumnWidth(0, 25)
        self.table_sessions.setColumnWidth(1, 80)
        self.table_sessions.setColumnWidth(2, 85)
        self.table_sessions.setColumnWidth(3, 100)
        self.table_sessions.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table_sessions.setColumnWidth(4, 100)
        self.table_sessions.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.table_sessions.setColumnWidth(5, 80)
        self.table_sessions.setColumnWidth(6, 35)
        self.table_sessions.verticalHeader().setVisible(False)
        self.table_sessions.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_sessions.setStyleSheet("""
            QTableWidget {
                border: 1px solid #374151;
                background-color: #111827;
                color: #e5e7eb;
            }
            QHeaderView::section {
                background-color: #1f2937;
                color: #9ca3af;
                padding: 4px;
                border: 1px solid #374151;
                font-weight: bold;
            }
        """)
        batch_layout.addWidget(self.table_sessions)
        
        btn_session_control_layout = QHBoxLayout()
        self.btn_add_session = QPushButton("➕ Thêm")
        self.btn_delete_session = QPushButton("🗑️ Xóa")
        self.btn_import_batch = QPushButton("📂 Nhập Excel")
        
        for btn in [self.btn_add_session, self.btn_delete_session, self.btn_import_batch]:
            btn.setStyleSheet("""
                QPushButton {
                    background-color: transparent; 
                    border: 1px solid #4b5563; 
                    padding: 4px 8px; 
                    color: #d1d5db; 
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #374151;
                    color: white;
                }
                QPushButton:pressed {
                    background-color: #4b5563;
                }
            """)
            btn_session_control_layout.addWidget(btn)
        btn_session_control_layout.addStretch()
        batch_layout.addLayout(btn_session_control_layout)
        
        left_layout.addLayout(batch_layout)
        
        # Queue Buttons
        queue_layout = QHBoxLayout()
        self.btn_add_queue = QPushButton("+ Thêm vào hàng chờ")
        self.btn_add_queue.setStyleSheet("""
            QPushButton {
                background-color: transparent; 
                border: 1px solid #8b5cf6; 
                color: #8b5cf6; 
                padding: 8px; 
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #8b5cf6;
                color: white;
            }
            QPushButton:pressed {
                background-color: #7c3aed;
                border-color: #7c3aed;
            }
        """)
        self.btn_manage_queue = QPushButton("📋 Quản lý hàng chờ (0)")
        self.btn_manage_queue.setStyleSheet("""
            QPushButton {
                background-color: #1f2937; 
                color: white; 
                padding: 8px; 
                border-radius: 5px;
                border: 1px solid #374151;
            }
            QPushButton:hover {
                background-color: #374151;
                border-color: #4b5563;
            }
            QPushButton:pressed {
                background-color: #111827;
            }
        """)
        queue_layout.addWidget(self.btn_add_queue)
        queue_layout.addWidget(self.btn_manage_queue)
        left_layout.addLayout(queue_layout)
        
        # Action Buttons
        action_layout = QHBoxLayout()
        self.btn_run = QPushButton("🚀 CHẠY NGAY")
        self.btn_run.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71; 
                color: white; 
                font-weight: bold; 
                padding: 12px; 
                border-radius: 5px;
                border: none;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
            QPushButton:pressed {
                background-color: #1e8449;
            }
        """)
        self.btn_pause = QPushButton("⏸ TẠM DỪNG")
        self.btn_pause.setEnabled(False)
        self.btn_pause.setStyleSheet("""
            QPushButton {
                background-color: #d1d5db; 
                color: #4b5563; 
                font-weight: bold; 
                padding: 12px; 
                border-radius: 5px;
                border: none;
            }
        """)
        self.btn_stop = QPushButton("⏹ DỪNG")
        self.btn_stop.setEnabled(False)
        self.btn_stop.setStyleSheet("""
            QPushButton {
                background-color: #d1d5db; 
                color: #4b5563; 
                font-weight: bold; 
                padding: 12px; 
                border-radius: 5px;
                border: none;
            }
        """)
        
        action_layout.addWidget(self.btn_run)
        action_layout.addWidget(self.btn_pause)
        action_layout.addWidget(self.btn_stop)
        left_layout.addLayout(action_layout)
        
        # Progress Stats Panel
        stats_group = QGroupBox("Tiến trình")
        stats_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #374151;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: #e5e7eb;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
        """)
        stats_layout = QVBoxLayout(stats_group)
        stats_layout.setSpacing(6)
        
        # Processing: X/Y
        self.lbl_processing = QLabel("⏳ Đang xử lý: 0/0")
        self.lbl_processing.setStyleSheet("""
            QLabel {
                color: #60a5fa;
                font-size: 14px;
                font-weight: bold;
                padding: 4px 8px;
                background-color: rgba(96, 165, 250, 0.1);
                border-radius: 4px;
            }
        """)
        stats_layout.addWidget(self.lbl_processing)
        
        # Success count
        self.lbl_success = QLabel("✅ Thành công: 0")
        self.lbl_success.setStyleSheet("""
            QLabel {
                color: #34d399;
                font-size: 14px;
                font-weight: bold;
                padding: 4px 8px;
                background-color: rgba(52, 211, 153, 0.1);
                border-radius: 4px;
            }
        """)
        stats_layout.addWidget(self.lbl_success)
        
        # Failure count
        self.lbl_failure = QLabel("❌ Thất bại: 0")
        self.lbl_failure.setStyleSheet("""
            QLabel {
                color: #f87171;
                font-size: 14px;
                font-weight: bold;
                padding: 4px 8px;
                background-color: rgba(248, 113, 113, 0.1);
                border-radius: 4px;
            }
        """)
        stats_layout.addWidget(self.lbl_failure)
        
        left_layout.addWidget(stats_group)
        
        # Initialize stats counters
        self.stats_total = 0
        self.stats_processed = 0
        self.stats_success = 0
        self.stats_failure = 0
        
        left_layout.addStretch()
        
        # Right Panel: Table
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(10, 0, 0, 0)
        
        # Filter Layout above table
        filter_layout = QHBoxLayout()
        self.chk_select_all = QCheckBox("Tích tất cả")
        self.chk_select_all.stateChanged.connect(self.toggle_select_all)
        filter_layout.addWidget(self.chk_select_all)
        filter_layout.addStretch()
        
        filter_layout.addWidget(QLabel("Xem phiên:"))
        self.combo_session_filter = QComboBox()
        self.combo_session_filter.setMinimumWidth(120)
        self.combo_session_filter.addItem("Tất cả")
        self.combo_session_filter.currentTextChanged.connect(self.on_session_filter_changed)
        filter_layout.addWidget(self.combo_session_filter)
        
        self.lbl_progress_header = QLabel("Tiến độ:")
        filter_layout.addWidget(self.lbl_progress_header)
        self.combo_filter = QComboBox()
        self.combo_filter.addItems(["Tất cả", "Đang chọn", "Đang chờ", "Đang tạo", "Hoàn thành", "Lỗi"])
        self.combo_filter.currentTextChanged.connect(self.filter_table)
        filter_layout.addWidget(self.combo_filter)
        right_layout.addLayout(filter_layout)
        
        self.table_tasks = QTableWidget(0, 6)
        
        # Column widths
        self.table_tasks.setHorizontalHeaderLabels([" ", "STT", "Anh tham chieu", "Prompt", "Ket qua", "Tien do"])
        self.table_tasks.setColumnWidth(0, 30)
        self.table_tasks.setColumnWidth(1, 40)
        self.table_tasks.setColumnWidth(2, 150)
        self.table_tasks.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table_tasks.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.table_tasks.setColumnWidth(5, 130)
        
        self.table_tasks.verticalHeader().setVisible(False)
        self.table_tasks.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        right_layout.addWidget(self.table_tasks)
        
        # Bottom Toolbar
        toolbar_layout = QHBoxLayout()
        
        # Nút Đóng Chrome màu đỏ nổi bật
        self.btn_tb_close_chrome = QPushButton("🛑 Đóng Chrome")
        self.btn_tb_close_chrome.setStyleSheet("""
            QPushButton {
                background-color: #ef4444; 
                border: 1px solid #dc2626; 
                padding: 5px 12px; 
                color: white; 
                font-weight: bold;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #dc2626;
                border-color: #b91c1c;
            }
            QPushButton:pressed {
                background-color: #b91c1c;
                border-color: #991b1b;
            }
        """)
        toolbar_layout.addWidget(self.btn_tb_close_chrome)
        
        self.btn_tb_delete = QPushButton("🗑️ Xóa")
        self.btn_tb_delete_all = QPushButton("🧹 Xóa hết")
        self.btn_tb_rerun_err = QPushButton("🔄 Chạy lại lỗi")
        self.btn_tb_run_sel = QPushButton("▶️ Chạy mục chọn")
        self.btn_tb_load_real_data = QPushButton("📡 Load dữ liệu thực tế")
        
        for btn in [self.btn_tb_delete, self.btn_tb_delete_all, 
                    self.btn_tb_rerun_err, self.btn_tb_run_sel, self.btn_tb_load_real_data]:
            btn.setStyleSheet("""
                QPushButton {
                    background-color: transparent; 
                    border: 1px solid #4b5563; 
                    padding: 5px 10px; 
                    color: #d1d5db; 
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #374151;
                    color: white;
                }
                QPushButton:pressed {
                    background-color: #4b5563;
                }
            """)
            toolbar_layout.addWidget(btn)
            
        toolbar_layout.addStretch()

        self.btn_previous_page = QPushButton("< Trang trước")
        self.lbl_page_info = QLabel("Trang 1/1")
        self.btn_next_page = QPushButton("Trang sau >")
        self.btn_previous_page.setEnabled(False)
        self.btn_next_page.setEnabled(False)
        toolbar_layout.addWidget(self.btn_previous_page)
        toolbar_layout.addWidget(self.lbl_page_info)
        toolbar_layout.addWidget(self.btn_next_page)
        right_layout.addLayout(toolbar_layout)
        
        self.splitter.addWidget(left_panel)
        self.splitter.addWidget(right_panel)
        self.splitter.setSizes([450, 650])
        
        # Connect signals
        self.btn_run.clicked.connect(self.start_tasks)
        self.btn_add_queue.clicked.connect(self.add_to_queue)
        self.btn_pause.clicked.connect(self.pause_tasks)
        self.btn_stop.clicked.connect(self.stop_tasks)
        self.table_tasks.itemChanged.connect(self.on_item_changed)
        self.btn_tb_close_chrome.clicked.connect(self.close_chrome)
        self.btn_tb_delete.clicked.connect(self.delete_selected_tasks)
        self.btn_tb_delete_all.clicked.connect(self.delete_all_tasks)
        self.btn_tb_rerun_err.clicked.connect(self.rerun_failed_tasks)
        self.btn_tb_run_sel.clicked.connect(self.run_selected_tasks)
        self.btn_tb_load_real_data.clicked.connect(self.load_real_data)
        self.btn_previous_page.clicked.connect(self.previous_page)
        self.btn_next_page.clicked.connect(self.next_page)
        
        # Connect batch session signals
        self.btn_add_session.clicked.connect(self.add_session_row)
        self.btn_delete_session.clicked.connect(self.delete_selected_sessions)
        self.btn_import_batch.clicked.connect(self.import_batch_sessions_config)
        self.table_sessions.itemChanged.connect(self.on_session_item_changed)
        
        self.current_session_task_ids = []
        
        # Tự động lưu cấu hình khi thay đổi
        self.combo_model.currentTextChanged.connect(self.save_config)
        self.chk_720p.stateChanged.connect(self.save_config)
        self.chk_1080p.stateChanged.connect(self.save_config)
        self.chk_4k.stateChanged.connect(self.save_config)
        self.combo_ratio.currentTextChanged.connect(self.save_config)
        self.spin_images_per_prompt.valueChanged.connect(self.save_config)
        self.spin_threads.valueChanged.connect(self.save_config)
        self.spin_delay_min.valueChanged.connect(self.save_config)
        self.spin_delay_max.valueChanged.connect(self.save_config)
        self.combo_ref_mode.currentTextChanged.connect(self.save_config)
        self.chk_lock_seed.stateChanged.connect(self.save_config)
        self.line_seed.textChanged.connect(self.save_config)
        
        self.workers = []
        self.task_queue = []
        self.active_workers_count = 0
        self.last_start_time = 0
        self.queue_timer = QTimer(self)
        self.queue_timer.timeout.connect(self.process_queue)
        
        self.hourglass_icon = "⏳"
        self.anim_timer = QTimer(self)
        self.anim_timer.timeout.connect(self.animate_hourglass)
        
        self.reset_running_tasks_to_pending()
        QTimer.singleShot(50, self.load_tasks)
        QTimer.singleShot(100, self.load_config)
        QTimer.singleShot(150, self.load_sessions)
 
    def browse_folder(self, line_edit):
        folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục")
        if folder:
            line_edit.setText(folder)

    def load_sessions(self):
        self.table_sessions.blockSignals(True)
        self.table_sessions.setRowCount(0)
        
        db = SessionLocal()
        sessions = db.query(VideoSession).order_by(VideoSession.id.asc()).all()
        
        # Nếu chưa có phiên nào, tạo 6 phiên mặc định
        if not sessions:
            for i in range(1, 7):
                new_sess = VideoSession(name=f"Phiên {i}", ref_dir="", save_dir="", prompts_text="", status="PENDING")
                db.add(new_sess)
            db.commit()
            sessions = db.query(VideoSession).order_by(VideoSession.id.asc()).all()
            
        db.close()
        
        for sess in sessions:
            self.add_session_row_to_table(sess)
            
        self.table_sessions.blockSignals(False)
        self.update_session_filter_combobox()

    def update_session_filter_combobox(self):
        self.combo_session_filter.blockSignals(True)
        current_text = self.combo_session_filter.currentText()
        self.combo_session_filter.clear()
        self.combo_session_filter.addItem("Tất cả")
        
        db = SessionLocal()
        sessions = db.query(VideoSession).order_by(VideoSession.id.asc()).all()
        for sess in sessions:
            self.combo_session_filter.addItem(sess.name, sess.id)
        db.close()
        
        idx = self.combo_session_filter.findText(current_text)
        if idx >= 0:
            self.combo_session_filter.setCurrentIndex(idx)
        else:
            self.combo_session_filter.setCurrentIndex(0)
        self.combo_session_filter.blockSignals(False)

    def add_session_row_to_table(self, sess):
        row = self.table_sessions.rowCount()
        self.table_sessions.insertRow(row)
        
        # Cột 0: Checkbox
        chk_widget = QWidget()
        chk_layout = QHBoxLayout(chk_widget)
        chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        chk_layout.setContentsMargins(0, 0, 0, 0)
        chk = QCheckBox()
        chk.setChecked(True)
        chk_layout.addWidget(chk)
        self.table_sessions.setCellWidget(row, 0, chk_widget)
        
        # Cột 1: Tên phiên
        name_item = QTableWidgetItem(sess.name)
        name_item.setData(Qt.ItemDataRole.UserRole, sess.id)
        self.table_sessions.setItem(row, 1, name_item)
        
        # Cột 2: Nút Prompts
        prompts = [p.strip() for p in (sess.prompts_text or "").split('\n') if p.strip()]
        btn_prompts = QPushButton(f"📝 Prompt ({len(prompts)})")
        btn_prompts.setStyleSheet("""
            QPushButton {
                background-color: #374151; 
                color: #e5e7eb; 
                border: 1px solid #4b5563; 
                border-radius: 4px;
                padding: 2px;
            }
            QPushButton:hover {
                background-color: #4b5563;
            }
        """)
        btn_prompts.clicked.connect(lambda _, s_id=sess.id: self.edit_session_prompts_by_id(s_id))
        self.table_sessions.setCellWidget(row, 2, btn_prompts)
        
        # Cột 3: Folder tham chiếu
        ref_widget = QWidget()
        ref_layout = QHBoxLayout(ref_widget)
        ref_layout.setContentsMargins(2, 2, 2, 2)
        ref_layout.setSpacing(2)
        ref_line = QLineEdit(sess.ref_dir or "")
        ref_line.setPlaceholderText("Folder...")
        ref_line.setReadOnly(True)
        ref_btn = QPushButton("📁")
        ref_btn.setFixedSize(24, 24)
        ref_btn.clicked.connect(lambda _, s_id=sess.id, le=ref_line: self.browse_session_folder_by_id(s_id, "ref", le))
        ref_layout.addWidget(ref_line)
        ref_layout.addWidget(ref_btn)
        self.table_sessions.setCellWidget(row, 3, ref_widget)
        
        # Cột 4: Folder lưu
        save_widget = QWidget()
        save_layout = QHBoxLayout(save_widget)
        save_layout.setContentsMargins(2, 2, 2, 2)
        save_layout.setSpacing(2)
        save_line = QLineEdit(sess.save_dir or "")
        save_line.setPlaceholderText("Folder...")
        save_line.setReadOnly(True)
        save_btn = QPushButton("📁")
        save_btn.setFixedSize(24, 24)
        save_btn.clicked.connect(lambda _, s_id=sess.id, le=save_line: self.browse_session_folder_by_id(s_id, "save", le))
        save_layout.addWidget(save_line)
        save_layout.addWidget(save_btn)
        self.table_sessions.setCellWidget(row, 4, save_widget)
        
        # Cột 5: Trạng thái
        status_item = QTableWidgetItem(sess.status or "PENDING")
        status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        status_item.setFlags(status_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        self.table_sessions.setItem(row, 5, status_item)
        
        # Cột 6: Hành động (Chạy lại phiên)
        btn_run = QPushButton("🔄")
        btn_run.setToolTip("Chạy lại riêng phiên này")
        btn_run.setStyleSheet("""
            QPushButton {
                background-color: transparent; 
                border: none;
                color: #60a5fa;
                font-weight: bold;
            }
            QPushButton:hover {
                color: #3b82f6;
            }
        """)
        btn_run.clicked.connect(lambda _, s_id=sess.id: self.rerun_session_by_id(s_id))
        self.table_sessions.setCellWidget(row, 6, btn_run)

    def add_session_row(self):
        db = SessionLocal()
        count = db.query(VideoSession).count()
        new_sess = VideoSession(name=f"Phiên {count + 1}", ref_dir="", save_dir="", prompts_text="", status="PENDING")
        db.add(new_sess)
        db.commit()
        
        self.add_session_row_to_table(new_sess)
        db.close()
        self.update_session_filter_combobox()
        logging.info(f"[UI] Đã thêm phiên mới: {new_sess.name}")

    def delete_selected_sessions(self):
        selected_ranges = self.table_sessions.selectedRanges()
        if not selected_ranges:
            QMessageBox.warning(self, "Thông báo", "Vui lòng chọn hàng cần xóa trên bảng!")
            return
            
        rows_to_delete = []
        for r in selected_ranges:
            for row in range(r.topRow(), r.bottomRow() + 1):
                rows_to_delete.append(row)
                
        rows_to_delete = sorted(list(set(rows_to_delete)), reverse=True)
        
        db = SessionLocal()
        for row in rows_to_delete:
            item = self.table_sessions.item(row, 1)
            if item:
                sess_id = item.data(Qt.ItemDataRole.UserRole)
                if sess_id:
                    sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
                    if sess:
                        # Xóa các task con liên kết trước
                        db.query(Task).filter(Task.session_id == sess_id, Task.task_type == "video").delete()
                        db.delete(sess)
            self.table_sessions.removeRow(row)
            
        db.commit()
        db.close()
        self.update_session_filter_combobox()
        logging.info("[UI] Đã xóa các phiên được chọn.")

    def edit_session_prompts_by_id(self, sess_id):
        db = SessionLocal()
        sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
        if not sess:
            db.close()
            return
            
        initial_text = sess.prompts_text or ""
        dialog = PromptEditDialog(self, sess.name, initial_text)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            imported_files = dialog.get_imported_prompt_files()
            session_ids_to_sync = []

            if imported_files:
                original_ref_dir = sess.ref_dir or ""
                for index, (file_path, prompts_text) in enumerate(imported_files):
                    session_name = os.path.splitext(os.path.basename(file_path))[0]
                    save_dir = os.path.splitext(os.path.abspath(file_path))[0]

                    if index == 0:
                        imported_session = sess
                        imported_session.name = session_name
                        imported_session.save_dir = save_dir
                        imported_session.prompts_text = prompts_text
                        imported_session.status = "PENDING"
                    else:
                        imported_session = VideoSession(
                            name=session_name,
                            ref_dir=original_ref_dir,
                            save_dir=save_dir,
                            prompts_text=prompts_text,
                            status="PENDING"
                        )
                        db.add(imported_session)

                    db.flush()
                    session_ids_to_sync.append(imported_session.id)
            else:
                sess.prompts_text = dialog.get_text()
                session_ids_to_sync.append(sess_id)

            db.commit()
            db.close()

            for imported_session_id in session_ids_to_sync:
                self.sync_session_tasks(imported_session_id, "video")
            self.load_sessions()
            self.load_tasks()
            return
        db.close()

    def browse_session_folder_by_id(self, sess_id, col_type, line_edit):
        folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục")
        if folder:
            line_edit.setText(folder)
            
            db = SessionLocal()
            sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
            if sess:
                if col_type == "ref":
                    sess.ref_dir = folder
                elif col_type == "save":
                    sess.save_dir = folder
                db.commit()
            db.close()

    def on_session_item_changed(self, item):
        if item.column() == 1:
            sess_id = item.data(Qt.ItemDataRole.UserRole)
            if sess_id:
                db = SessionLocal()
                sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
                if sess:
                    sess.name = item.text().strip()
                    db.commit()
                db.close()
                self.update_session_filter_combobox()

    def import_batch_sessions_config(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Chọn tệp cấu hình loạt", "", "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return
            
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            sheet = wb.active
            
            imported_count = 0
            db = SessionLocal()
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                    
                sess_name = str(row[0]).strip()
                ref_dir = str(row[1]).strip() if row[1] else ""
                save_dir = str(row[2]).strip() if row[2] else ""
                prompts_val = str(row[3]).strip() if row[3] else ""
                
                prompts_text = ""
                if prompts_val:
                    if os.path.exists(prompts_val) and prompts_val.endswith('.txt'):
                        try:
                            with open(prompts_val, 'r', encoding='utf-8') as pf:
                                prompts_text = pf.read()
                        except Exception as fe:
                            logging.warning(f"[UI] Không đọc được file prompt {prompts_val}: {fe}")
                    else:
                        prompts_text = prompts_val
                
                new_sess = VideoSession(
                    name=sess_name,
                    ref_dir=ref_dir,
                    save_dir=save_dir,
                    prompts_text=prompts_text,
                    status="PENDING"
                )
                db.add(new_sess)
                imported_count += 1
                
            db.commit()
            db.close()
            
            self.load_sessions()
            QMessageBox.information(self, "Thành công", f"Đã nhập thành công {imported_count} phiên từ file Excel!")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi đọc file cấu hình loạt: {str(e)}")

    def browse_folder(self, line_edit):
        folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục")
        if folder:
            line_edit.setText(folder)

    def create_ref_widget(self, ref_image_str=None):
        w = QWidget()
        l = QHBoxLayout(w)
        l.setContentsMargins(2, 2, 2, 2)
        l.setSpacing(2)
        
        ref_paths = ref_image_str.split(",") if ref_image_str else []
        for idx in range(3):
            btn = QPushButton()
            btn.setFixedSize(30, 30)
            btn.clicked.connect(lambda _, b=btn: self.on_ref_btn_clicked(b))
            
            if idx < len(ref_paths) and os.path.exists(ref_paths[idx]):
                btn.setText("📸")
                btn.setToolTip(ref_paths[idx])
                btn.setStyleSheet("background-color: #374151; border: 1px solid #4b5563; color: white;")
            else:
                text = "+7" if idx == 2 else "+"
                btn.setText(text)
                btn.setStyleSheet("background-color: transparent; border: 1px dashed #6b7280; color: #9ca3af;")
            l.addWidget(btn)
        return w
        
    def on_ref_btn_clicked(self, btn):
        file, _ = QFileDialog.getOpenFileName(self, "Chọn ảnh tham chiếu", "", "Images (*.png *.jpg *.jpeg)")
        if file:
            btn.setText("📸")
            btn.setToolTip(file)
            btn.setStyleSheet("background-color: #374151; border: 1px solid #4b5563; color: white;")
            
            widget = btn.parentWidget()
            if widget:
                for r in range(self.table_tasks.rowCount()):
                    if self.table_tasks.cellWidget(r, 2) == widget:
                        task_id = self.table_tasks.item(r, 1).data(Qt.ItemDataRole.UserRole)
                        if task_id:
                            buttons = widget.findChildren(QPushButton)
                            ref_paths = []
                            for b in buttons:
                                if b.toolTip() and b.toolTip() != b.text():
                                    ref_paths.append(b.toolTip())
                            
                            db = SessionLocal()
                            task = db.query(Task).filter(Task.id == task_id).first()
                            if task:
                                task.reference_image = ",".join(ref_paths) if ref_paths else None
                                db.commit()
                                logging.info(f"[UI] Đã lưu ảnh tham chiếu cho Task ID {task_id}: {task.reference_image}")
                            db.close()
                        break

    def open_path(self, path):
        if path and os.path.exists(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(path)))

    def open_containing_folder(self, path):
        if path and os.path.exists(path):
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(os.path.abspath(path))))

    def create_result_widget(self, result_path):
        if not result_path:
            return None

        # Check if result file is a video
        ext = os.path.splitext(result_path.lower())[1] if result_path else ""
        is_video = ext in ['.mp4', '.avi', '.mov', '.mkv', '.webm']

        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(4, 2, 4, 2)
        layout.setSpacing(6)

        btn = QPushButton()
        btn.setFixedSize(72, 54)
        btn.setToolTip(result_path)
        btn.clicked.connect(lambda _, p=result_path: self.open_path(p))
        btn.setStyleSheet("""
            QPushButton {
                background-color: #111827;
                border: 1px solid #374151;
                border-radius: 4px;
                color: #e5e7eb;
            }
            QPushButton:hover {
                border-color: #60a5fa;
            }
        """)

        if is_video:
            btn.setText("🎥")
        else:
            # Check image cache
            if result_path in self.thumbnail_cache:
                pixmap = self.thumbnail_cache[result_path]
                if not pixmap.isNull():
                    btn.setIcon(QIcon(pixmap))
                    btn.setIconSize(btn.size())
                else:
                    btn.setText(os.path.basename(result_path))
            else:
                btn.setText("⏳")
                # Start loading background image if the file actually exists
                if os.path.exists(result_path):
                    if result_path not in self.loading_thumbnails:
                        self.loading_thumbnails[result_path] = []
                        self.queue_thumbnail_load(result_path)
                    self.loading_thumbnails[result_path].append(btn)
                else:
                    btn.setText(os.path.basename(result_path))

        layout.addWidget(btn)

        label = QLabel(os.path.basename(result_path))
        label.setToolTip(result_path)
        label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        layout.addWidget(label, 1)
        return widget

    def queue_thumbnail_load(self, path):
        from ui.components.thumbnail_loader import ImageLoaderRunnable
        from PyQt6.QtCore import QThreadPool
        
        runnable = ImageLoaderRunnable(path, 64, 46)
        runnable.signals.loaded.connect(self.on_thumbnail_loaded)
        runnable.signals.failed.connect(self.on_thumbnail_failed)
        QThreadPool.globalInstance().start(runnable)

    def on_thumbnail_loaded(self, path, image):
        from PyQt6.QtGui import QPixmap
        pixmap = QPixmap.fromImage(image)
        self.thumbnail_cache[path] = pixmap
        
        buttons = self.loading_thumbnails.pop(path, [])
        for btn in buttons:
            try:
                btn.setText("")
                btn.setIcon(QIcon(pixmap))
                btn.setIconSize(btn.size())
            except RuntimeError:
                pass # Widget was deleted

    def on_thumbnail_failed(self, path):
        from PyQt6.QtGui import QPixmap
        self.thumbnail_cache[path] = QPixmap() # Cache empty to prevent retry
        
        buttons = self.loading_thumbnails.pop(path, [])
        for btn in buttons:
            try:
                btn.setText(os.path.basename(path))
            except RuntimeError:
                pass

    def create_status_widget(self, status, result_path=None):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(4, 2, 4, 2)
        layout.setSpacing(6)

        label = QLabel(status)
        layout.addWidget(label, 1)

        if status == "COMPLETED" and result_path and os.path.exists(result_path):
            btn = QPushButton("📁")
            btn.setFixedSize(28, 28)
            btn.setToolTip("Mở thư mục chứa ảnh")
            btn.clicked.connect(lambda _, p=result_path: self.open_containing_folder(p))
            btn.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    border: 1px solid #4b5563;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #374151;
                }
            """)
            layout.addWidget(btn)
        return widget

    def on_session_filter_changed(self):
        self.current_page = 1
        self.load_tasks()

    def previous_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.load_tasks()

    def next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_tasks()

    def update_pagination_controls(self):
        self.lbl_page_info.setText(f"Trang {self.current_page}/{self.total_pages}")
        self.btn_previous_page.setEnabled(self.current_page > 1)
        self.btn_next_page.setEnabled(self.current_page < self.total_pages)

    def load_tasks(self):
        if hasattr(self, '_load_tasks_timer') and self._load_tasks_timer.isActive():
            self._load_tasks_timer.stop()
            
        self.loading_thumbnails.clear() # Clear pending thumbnail registrations

        self.table_tasks.blockSignals(True)
        self.table_tasks.setRowCount(0)
        
        db = SessionLocal()
        
        # Check if we should filter by session
        selected_sess_id = None
        if hasattr(self, 'combo_session_filter') and self.combo_session_filter.currentIndex() > 0:
            selected_sess_id = self.combo_session_filter.currentData()
            
        # Query tasks for this type
        query = db.query(Task).filter(Task.task_type == "video")
        if selected_sess_id is not None:
            query = query.filter(Task.session_id == selected_sess_id)

        total_count = query.count()
        completed_count = query.filter(Task.status == "COMPLETED").count()
        self.total_pages = max(1, (total_count + self.page_size - 1) // self.page_size)
        self.current_page = min(max(1, self.current_page), self.total_pages)
        page_offset = (self.current_page - 1) * self.page_size
        tasks = query.order_by(Task.id.asc()).offset(page_offset).limit(self.page_size).all()
        db.close()


        self.table_tasks.setUpdatesEnabled(False)
        for i, task in enumerate(tasks):
            row_idx = self.table_tasks.rowCount()
            self.table_tasks.insertRow(row_idx)

            chk_container = QWidget()
            chk_layout = QHBoxLayout(chk_container)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            chk = QCheckBox()
            chk_layout.addWidget(chk)
            self.table_tasks.setCellWidget(row_idx, 0, chk_container)

            stt_item = QTableWidgetItem(str(page_offset + row_idx + 1))
            stt_item.setData(Qt.ItemDataRole.UserRole, task.id)
            stt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            stt_item.setFlags(stt_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table_tasks.setItem(row_idx, 1, stt_item)

            self.table_tasks.setCellWidget(row_idx, 2, self.create_ref_widget(task.reference_image))

            prompt_item = QTableWidgetItem(task.prompt)
            prompt_item.setData(Qt.ItemDataRole.UserRole, task.id)
            self.table_tasks.setItem(row_idx, 3, prompt_item)

            result_item = QTableWidgetItem(task.result_path or "")
            result_item.setFlags(result_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table_tasks.setItem(row_idx, 4, result_item)
            result_widget = self.create_result_widget(task.result_path)
            if task.status == "COMPLETED" and result_widget:
                self.table_tasks.setCellWidget(row_idx, 4, result_widget)
                self.table_tasks.setRowHeight(row_idx, 64)

            status_item = QTableWidgetItem(task.status)
            status_item.setFlags(status_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table_tasks.setItem(row_idx, 5, status_item)
            if task.status == "COMPLETED":
                self.table_tasks.setCellWidget(row_idx, 5, self.create_status_widget(task.status, task.result_path))

        if total_count > 0:
            self.lbl_progress_header.setText(f"Tiến độ({completed_count}/{total_count}):")
        else:
            self.lbl_progress_header.setText("Tiến độ:")
        self.update_pagination_controls()
        self.table_tasks.setUpdatesEnabled(True)
        self.table_tasks.blockSignals(False)
        self.filter_table()
        
    def on_item_changed(self, item):
        if item.column() == 3:
            task_id = item.data(Qt.ItemDataRole.UserRole)
            if task_id:
                db = SessionLocal()
                task = db.query(Task).filter(Task.id == task_id).first()
                if task:
                    task.prompt = item.text().strip()
                    db.commit()
                db.close()
        
    def filter_table(self):
        filter_text = self.combo_filter.currentText()
        for i in range(self.table_tasks.rowCount()):
            if filter_text == "Tất cả":
                self.table_tasks.setRowHidden(i, False)
                continue
                
            if filter_text == "Đang chọn":
                widget = self.table_tasks.cellWidget(i, 0)
                if widget:
                    chk = widget.findChild(QCheckBox)
                    self.table_tasks.setRowHidden(i, not chk.isChecked())
                continue
                
            status_item = self.table_tasks.item(i, 5)
            if status_item:
                status = status_item.text()
                mapping = {"Đang chờ": "PENDING", "Đang tạo": "RUNNING", "Hoàn thành": "COMPLETED", "Lỗi": "ERROR"}
                target_status = mapping.get(filter_text, "")
                self.table_tasks.setRowHidden(i, target_status not in status)

    def add_to_queue(self):
        prompts = [p.strip() for p in prompts if p.strip()]
        if not prompts: return
        logging.info(f"[UI] Thêm {len(prompts)} prompts vào hàng chờ.")
        db = SessionLocal()
        for p in prompts:
            logging.info(f"[UI] Prompt thêm: '{p}'")
            db.add(Task(prompt=p, status="PENDING", task_type="video"))
        db.commit()
        db.close()
        self.load_tasks()

    def start_tasks(self):
        self.start_batch_sessions()

    def sync_session_tasks(self, sess_id, task_type="video"):
        db = SessionLocal()
        sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
        if not sess:
            db.close()
            return False
            
        prompts = [p.strip() for p in (sess.prompts_text or "").split('\n') if p.strip()]
        if not prompts:
            # If no prompts, we should probably clear existing tasks
            db.query(Task).filter(Task.session_id == sess_id, Task.task_type == task_type).delete()
            db.commit()
            db.close()
            return False
            
        import os
        save_path = sess.save_dir.strip() if sess.save_dir else ""
        if not save_path:
            src_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            save_path = os.path.join(src_dir, "output", sess.name.replace(" ", "_"))
            sess.save_dir = save_path
            db.commit()
            
        if not os.path.exists(save_path):
            try:
                os.makedirs(save_path)
            except Exception as e:
                import logging
                logging.warning(f"[UI] Không tạo được thư mục lưu {save_path}: {e}")
        
        ref_images = []
        if sess.ref_dir and os.path.exists(sess.ref_dir):
            try:
                for file in os.listdir(sess.ref_dir):
                    if file.lower().endswith(('.png', '.jpg', '.jpeg')):
                        ref_images.append(os.path.join(sess.ref_dir, file))
            except Exception as e:
                import logging
                logging.warning(f"[UI] Không quét được thư mục ảnh tham chiếu {sess.ref_dir}: {e}")
        
        ref_mode = self.combo_ref_mode.currentText()
        tasks = db.query(Task).filter(Task.session_id == sess_id, Task.task_type == task_type).order_by(Task.id.asc()).all()
        
        if len(tasks) != len(prompts):
            db.query(Task).filter(Task.session_id == sess_id, Task.task_type == task_type).delete()
            for idx, p in enumerate(prompts):
                ref_image_path = None
                if ref_images:
                    if ref_mode == "1 cho tất cả":
                        ref_image_path = ref_images[0]
                    else:
                        ref_image_path = ref_images[idx % len(ref_images)]
                        
                db.add(Task(
                    prompt=p,
                    reference_image=ref_image_path,
                    status="PENDING",
                    task_type=task_type,
                    session_id=sess_id
                ))
        else:
            for idx, t in enumerate(tasks):
                ref_image_path = None
                if ref_images:
                    if ref_mode == "1 cho tất cả":
                        ref_image_path = ref_images[0]
                    else:
                        ref_image_path = ref_images[idx % len(ref_images)]
                        
                if t.prompt != prompts[idx]:
                    t.prompt = prompts[idx]
                    t.status = "PENDING"
                    t.result_path = None
                    
                t.reference_image = ref_image_path
                
        db.commit()
        db.close()
        return True

    def start_batch_sessions(self):
        logging.info("[UI] Bắt đầu chạy tuần tự theo phiên.")
        
        db = SessionLocal()
        active_account = db.query(Account).filter(Account.is_active == True).order_by(Account.position.asc()).first()
        db.close()
        if not active_account:
            QMessageBox.warning(self, "Thiếu tài khoản", "Vui lòng thêm hoặc kích hoạt ít nhất một tài khoản ở tab 'Cài đặt hệ thống'!")
            return

        selected_sess_ids = []
        for r in range(self.table_sessions.rowCount()):
            chk_widget = self.table_sessions.cellWidget(r, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk and chk.isChecked():
                    sess_item = self.table_sessions.item(r, 1)
                    if sess_item:
                        sess_id = sess_item.data(Qt.ItemDataRole.UserRole)
                        if sess_id:
                            selected_sess_ids.append(sess_id)
                        
        if not selected_sess_ids:
            # Fallback: check if a row is selected (highlighted)
            selected_items = self.table_sessions.selectedItems()
            if selected_items:
                r = selected_items[0].row()
                sess_item = self.table_sessions.item(r, 1)
                if sess_item:
                    sess_id = sess_item.data(Qt.ItemDataRole.UserRole)
                    if sess_id:
                        selected_sess_ids.append(sess_id)
                        
        if not selected_sess_ids:
            QMessageBox.warning(self, "Thiếu dữ liệu", "Vui lòng tích chọn hoặc chọn một phiên trên bảng để chạy!")
            return

        db = SessionLocal()
        valid_sess_ids = []
        
        for sess_id in selected_sess_ids:
            if self.sync_session_tasks(sess_id, "video"):
                sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
                if sess:
                    sess.status = "PENDING"
                    db.commit()
                valid_sess_ids.append(sess_id)
            
        db.close()
        
        if not valid_sess_ids:
            QMessageBox.warning(self, "Thiếu dữ liệu", "Không có phiên nào có prompt hợp lệ để chạy!")
            return
            
        self.load_sessions()
        
        self.active_batch_session_ids = valid_sess_ids
        self.current_batch_session_idx = 0
        self.run_next_batch_session()

    def run_next_batch_session(self):
        if not hasattr(self, 'active_batch_session_ids') or self.current_batch_session_idx >= len(self.active_batch_session_ids):
            logging.info("[UI] Đã hoàn thành chạy tất cả các phiên.")
            self.set_running_buttons_disabled()
            
            db = SessionLocal()
            for r in range(self.table_sessions.rowCount()):
                sess_item = self.table_sessions.item(r, 1)
                if sess_item:
                    sess_id = sess_item.data(Qt.ItemDataRole.UserRole)
                    if sess_id in getattr(self, 'active_batch_session_ids', []):
                        sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
                        if sess and sess.status == "RUNNING":
                            err_count = db.query(Task).filter(Task.session_id == sess_id, Task.status.like("ERROR%")).count()
                            sess.status = "ERROR" if err_count > 0 else "COMPLETED"
            db.commit()
            db.close()
            
            self.load_sessions()
            QMessageBox.information(self, "Hoàn thành", "Đã chạy xong tất cả các phiên được cấu hình!")
            return
            
        sess_id = self.active_batch_session_ids[self.current_batch_session_idx]
        
        db = SessionLocal()
        sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
        if not sess:
            db.close()
            self.current_batch_session_idx += 1
            self.run_next_batch_session()
            return
            
        sess.status = "RUNNING"
        sess_save_dir = sess.save_dir
        db.commit()
        db.close()
        
        for r in range(self.table_sessions.rowCount()):
            item = self.table_sessions.item(r, 1)
            if item and item.data(Qt.ItemDataRole.UserRole) == sess_id:
                status_item = self.table_sessions.item(r, 5)
                if status_item:
                    status_item.setText("⚙️ Đang chạy")
                break
                
        self.combo_session_filter.blockSignals(True)
        idx = self.combo_session_filter.findData(sess_id)
        if idx >= 0:
            self.combo_session_filter.setCurrentIndex(idx)
        self.combo_session_filter.blockSignals(False)

        self.current_page = 1
        self.load_tasks()
        
        db = SessionLocal()
        current_sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
        all_tasks = db.query(Task).filter(Task.session_id == sess_id, Task.task_type == "video").order_by(Task.id.asc()).all()
        tasks = [t for t in all_tasks if t.status == "PENDING"]
        
        if not tasks:
            sess_name = current_sess.name if current_sess else f"ID {sess_id}"
            logging.info(f"[UI] Phiên '{sess_name}' không có task PENDING, tự động chuyển phiên kế tiếp.")
            err_count = sum(1 for t in all_tasks if t.status and "ERROR" in t.status.upper())
            if current_sess:
                current_sess.status = "ERROR" if err_count > 0 else "COMPLETED"
                db.commit()
            self.load_sessions()
            self.current_batch_session_idx += 1
            db.close()
            QTimer.singleShot(100, self.run_next_batch_session)
            return
            
        
        self.stats_total = len(tasks)
        self.stats_processed = 0
        self.stats_success = 0
        self.stats_failure = 0
        self.update_stats_display()
        
        task_id_to_stt = {t.id: str(idx + 1) for idx, t in enumerate(all_tasks)}
        
        for t in tasks:
            task_config = {
                "model": self.combo_model.currentText(),
                "quality": [q.text() for q in [self.chk_720p, self.chk_1080p, self.chk_4k] if q.isChecked()],
                "aspect_ratio": self.combo_ratio.currentText(),
                "images_per_prompt": self.spin_images_per_prompt.value(),
                "save_path": sess_save_dir
            }
            
            stt = task_id_to_stt[t.id]
            self.task_queue.append({
                'task_id': t.id,
                'final_name': stt,
                'target': "labs.google/video",
                'config': task_config
            })
        db.close()
        
        if not self.queue_timer.isActive():
            self.set_running_buttons_enabled()
            self.queue_timer.start(1000)
            self.process_queue()

    def rerun_session_by_id(self, sess_id):
        db = SessionLocal()
        sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
        if not sess:
            db.close()
            return
            
        err_tasks = db.query(Task).filter(Task.session_id == sess_id, Task.status.like("ERROR%")).all()
        
        if not err_tasks:
            reply = QMessageBox.question(
                self, 
                "Xác nhận chạy lại", 
                f"Phiên '{sess.name}' không có lỗi. Bạn có chắc chắn muốn chạy lại toàn bộ phiên?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                db.close()
                return
            mode = "all"
        else:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Chạy lại phiên")
            msg_box.setText(f"Phiên '{sess.name}' có {len(err_tasks)} ảnh bị lỗi. Bạn muốn:")
            btn_err = msg_box.addButton("Chỉ chạy lại ảnh lỗi", QMessageBox.ButtonRole.ActionRole)
            btn_all = msg_box.addButton("Chạy lại toàn bộ phiên", QMessageBox.ButtonRole.ActionRole)
            btn_cancel = msg_box.addButton("Hủy", QMessageBox.ButtonRole.RejectRole)
            
            msg_box.exec()
            
            if msg_box.clickedButton() == btn_err:
                mode = "error"
            elif msg_box.clickedButton() == btn_all:
                mode = "all"
            else:
                db.close()
                return
                
        if mode == "error":
            db.query(Task).filter(Task.session_id == sess_id, Task.status.like("ERROR%")).update({"status": "PENDING"}, synchronize_session=False)
        else:
            db.query(Task).filter(Task.session_id == sess_id).update({"status": "PENDING", "result_path": None}, synchronize_session=False)
            
        sess.status = "PENDING"
        db.commit()
        db.close()
        
        self.load_sessions()
        
        self.active_batch_session_ids = [sess_id]
        self.current_batch_session_idx = 0
        self.run_next_batch_session()

    def check_and_advance_batch_session(self, task_id):
        db = SessionLocal()
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task or not task.session_id:
            db.close()
            return
            
        sess_id = task.session_id
        
        if not hasattr(self, 'active_batch_session_ids') or sess_id not in self.active_batch_session_ids:
            db.close()
            return
            
        active_count = db.query(Task).filter(
            Task.session_id == sess_id,
            Task.task_type == "video",
            Task.status.in_(["PENDING", "RUNNING"])
        ).count()
        
        in_queue_count = sum(1 for t in self.task_queue if t['task_id'] == task_id)
        
        db.close()
        
        if active_count == 0 and in_queue_count == 0:
            db = SessionLocal()
            sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
            if sess:
                err_count = db.query(Task).filter(Task.session_id == sess_id, Task.status.like("ERROR%")).count()
                sess.status = "ERROR" if err_count > 0 else "COMPLETED"
                db.commit()
            db.close()
            
            self.load_sessions()
            
            if hasattr(self, 'current_batch_session_idx') and sess_id == self.active_batch_session_ids[self.current_batch_session_idx]:
                self.current_batch_session_idx += 1
                QTimer.singleShot(2000, self.run_next_batch_session)

    def set_running_buttons_enabled(self):
        """Kích hoạt nút TẠM DỪNG (cam) và DỪNG (đỏ) khi đang chạy."""
        if not self.anim_timer.isActive():
            self.anim_timer.start(500)
        self.btn_pause.setEnabled(True)
        self.btn_pause.setText("⏸ TẠM DỪNG")
        self.btn_pause.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b; 
                color: white; 
                font-weight: bold; 
                padding: 12px; 
                border-radius: 5px;
                border: none;
            }
            QPushButton:hover {
                background-color: #fbbf24;
            }
            QPushButton:pressed {
                background-color: #d97706;
            }
        """)
        self.btn_stop.setEnabled(True)
        self.btn_stop.setStyleSheet("""
            QPushButton {
                background-color: #ef4444; 
                color: white; 
                font-weight: bold; 
                padding: 12px; 
                border-radius: 5px;
                border: none;
            }
            QPushButton:hover {
                background-color: #f87171;
            }
            QPushButton:pressed {
                background-color: #dc2626;
            }
        """)

    def set_running_buttons_disabled(self):
        """Vô hiệu hóa nút TẠM DỪNG và DỪNG khi không chạy."""
        self.anim_timer.stop()
        self.hourglass_icon = "⏳"
        self.update_stats_display()
        self.btn_pause.setEnabled(False)
        self.btn_pause.setText("⏸ TẠM DỪNG")
        self.btn_pause.setStyleSheet("""
            QPushButton {
                background-color: #d1d5db; 
                color: #4b5563; 
                font-weight: bold; 
                padding: 12px; 
                border-radius: 5px;
                border: none;
            }
        """)
        self.btn_stop.setEnabled(False)
        self.btn_stop.setStyleSheet("""
            QPushButton {
                background-color: #d1d5db; 
                color: #4b5563; 
                font-weight: bold; 
                padding: 12px; 
                border-radius: 5px;
                border: none;
            }
        """)

    def process_queue(self):
        max_threads = self.spin_threads.value()
        
        # Cleanup finished workers
        self.workers = [w for w in self.workers if w.isRunning() or not w.isFinished()]
        self.active_workers_count = sum(1 for w in self.workers if w.isRunning())
        
        if not self.task_queue:
            if self.active_workers_count == 0:
                self.queue_timer.stop()
                self.set_running_buttons_disabled()
            return
            
        current_time = time.time()
        
        delay_min = self.spin_delay_min.value()
        delay_max = self.spin_delay_max.value()
        
        # Calculate random delay if min != max, else just min
        delay_s = random.randint(delay_min, delay_max) if delay_max > delay_min else delay_min
        
        if self.active_workers_count < max_threads:
            if current_time - self.last_start_time >= delay_s or self.active_workers_count == 0:
                task_info = self.task_queue.pop(0)
                
                worker_config = task_info.get('config', getattr(self, 'base_config', {}).copy()).copy()
                worker_config["final_name"] = task_info['final_name']
                
                # Phân bổ tài khoản thông minh (Load balancing)
                # 1. Lấy danh sách các tài khoản active từ DB
                db = SessionLocal()
                active_accounts = db.query(Account).filter(Account.is_active == True).order_by(Account.position.asc()).all()
                db.close()
                
                selected_account_id = None
                if active_accounts:
                    # 2. Đếm số lượng tasks đang chạy của từng tài khoản active
                    active_workers = [w for w in self.workers if w.isRunning()]
                    account_task_counts = {}
                    for acc in active_accounts:
                        account_task_counts[acc.id] = 0
                        
                    for w in active_workers:
                        if w.account_id in account_task_counts:
                            account_task_counts[w.account_id] += 1
                            
                    # 3. Chọn tài khoản có số lượng tasks đang chạy ít nhất
                    sorted_accounts = sorted(active_accounts, key=lambda acc: account_task_counts[acc.id])
                    selected_account = sorted_accounts[0]
                    selected_account_id = selected_account.id
                    logging.info(f"[Queue] Phân bổ tài khoản {selected_account.email} (đang chạy {account_task_counts[selected_account.id]} luồng) cho Task ID {task_info['task_id']}")
                
                worker = AutomationWorker(task_info['task_id'], task_info['target'], worker_config, account_id=selected_account_id)
                worker.progress.connect(self.update_task_status)
                worker.task_finished.connect(self.on_task_finished)
                worker.error.connect(self.on_task_error)
                
                self.workers.append(worker)
                worker.start()
                
                self.active_workers_count += 1
                self.last_start_time = time.time()

    def pause_tasks(self):
        if "TẠM DỪNG" in self.btn_pause.text():
            logging.info("[UI] Tạm dừng các tasks.")
            self.anim_timer.stop()
            self.hourglass_icon = "⏳"
            self.update_stats_display()
            if getattr(self, 'queue_timer', None) and self.queue_timer.isActive():
                self.queue_timer.stop()
            for worker in self.workers:
                if worker.isRunning():
                    worker.pause()
            self.btn_pause.setText("▶ TIẾP TỤC")
            self.btn_pause.setStyleSheet("""
                QPushButton {
                    background-color: #2ecc71; 
                    color: white; 
                    font-weight: bold; 
                    padding: 12px; 
                    border-radius: 5px;
                    border: none;
                }
                QPushButton:hover {
                    background-color: #27ae60;
                }
                QPushButton:pressed {
                    background-color: #1e8449;
                }
            """)
        else:
            logging.info("[UI] Tiếp tục chạy các tasks.")
            self.anim_timer.start(500)
            if getattr(self, 'queue_timer', None):
                self.queue_timer.start(1000)
            for worker in self.workers:
                if worker.isRunning():
                    worker.resume()
            self.btn_pause.setText("⏸ TẠM DỪNG")
            self.btn_pause.setStyleSheet("""
                QPushButton {
                    background-color: #f59e0b; 
                    color: white; 
                    font-weight: bold; 
                    padding: 12px; 
                    border-radius: 5px;
                    border: none;
                }
                QPushButton:hover {
                    background-color: #fbbf24;
                }
                QPushButton:pressed {
                    background-color: #d97706;
                }
            """)

    def stop_tasks(self):
        logging.info("[UI] Dừng toàn bộ các tasks và xóa hàng chờ.")
        if getattr(self, 'queue_timer', None) and self.queue_timer.isActive():
            self.queue_timer.stop()
        self.task_queue.clear()
        for worker in self.workers:
            if worker.isRunning():
                worker.stop()
        self.reset_running_tasks_to_pending()
        self.set_running_buttons_disabled()

    def reset_running_tasks_to_pending(self):
        db = SessionLocal()
        try:
            db.query(Task).filter(Task.status == "RUNNING", Task.task_type == "video").update({"status": "PENDING"}, synchronize_session=False)
            db.commit()
        finally:
            db.close()
        for i in range(self.table_tasks.rowCount()):
            status_item = self.table_tasks.item(i, 5)
            if status_item and "RUNNING" in status_item.text():
                item = self.table_tasks.item(i, 1)
                if item:
                    self.update_table_row(item.data(Qt.ItemDataRole.UserRole), 5, "PENDING")
        active_task_ids = {worker.task_id for worker in self.workers if worker.isRunning()}
        active_task_ids.update(task_info.get('task_id') for task_info in self.task_queue)
        for task_id in active_task_ids:
            if task_id:
                self.update_table_row(task_id, 5, "PENDING")

    def shutdown_tasks(self):
        if getattr(self, 'queue_timer', None) and self.queue_timer.isActive():
            self.queue_timer.stop()
        self.task_queue.clear()
        for worker in self.workers:
            if worker.isRunning():
                worker.stop()
        self.reset_running_tasks_to_pending()

    def close_chrome(self):
        running = False
        if self.active_workers_count > 0 or self.task_queue:
            running = True
            
        if running:
            reply = QMessageBox.question(
                self, 
                "Xác nhận", 
                "Các luồng đang chạy dở. Bạn có chắc chắn muốn dừng tác vụ và đóng Chrome?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
                
        # Stop tasks (stops workers and resets status in DB)
        self.stop_tasks()
        
        # Kill Chrome processes
        try:
            from core.browser_manager import kill_all_registered_chromes
            kill_all_registered_chromes()
        except Exception as e:
            logging.error(f"[UI] Lỗi khi đóng Chrome: {e}")
            
        QMessageBox.information(self, "Thông báo", "Đã đóng tất cả trình duyệt Chrome do Bot mở.")

    def closeEvent(self, event):
        self.shutdown_tasks()
        super().closeEvent(event)

    def update_task_status(self, task_id, status):
        status_lower = status.lower()
        if "dừng" in status_lower or "dá»«ng" in status_lower or "stopped" in status_lower:
            status = "PENDING"
        elif status.startswith("Đang") or status.startswith("⏳") or status.startswith("⌛"):
            clean_status = status.lstrip("⏳⌛ ").strip()
            status = f"{self.hourglass_icon} {clean_status}"
        self.update_table_row(task_id, 5, status)

    def on_task_finished(self, task_id, result_path):
        self.update_table_row(task_id, 4, result_path)
        self.update_table_row(task_id, 5, "COMPLETED")
        self.stats_processed += 1
        self.stats_success += 1
        self.update_stats_display()
        self.check_and_advance_batch_session(task_id)

    def on_task_error(self, task_id, error_msg):
        self.update_table_row(task_id, 5, f"ERROR: {error_msg}")
        self.stats_processed += 1
        self.stats_failure += 1
        self.update_stats_display()
        self.check_and_advance_batch_session(task_id)
        
    def animate_hourglass(self):
        frames = ["⏳", "⌛"]
        try:
            idx = frames.index(self.hourglass_icon)
            self.hourglass_icon = frames[(idx + 1) % len(frames)]
        except ValueError:
            self.hourglass_icon = frames[0]
            
        self.update_stats_display()
        
        # Cập nhật biểu tượng đồng hồ cát động cho các dòng đang chạy trong bảng
        for i in range(self.table_tasks.rowCount()):
            status_item = self.table_tasks.item(i, 5)
            if status_item:
                text = status_item.text()
                if any(text.startswith(char) for char in frames) or text.startswith("Đang"):
                    clean_text = text.lstrip("⏳⌛🔃🔄 ").strip()
                    status_item.setText(f"{self.hourglass_icon} {clean_text}")

    def update_stats_display(self):
        self.lbl_processing.setText(f"{self.hourglass_icon} Đang xử lý: {self.stats_processed}/{self.stats_total}")
        self.lbl_success.setText(f"✅ Thành công: {self.stats_success}")
        self.lbl_failure.setText(f"❌ Thất bại: {self.stats_failure}")
        # Cập nhật label tiến độ trên filter bar
        self.lbl_progress_header.setText(f"Tiến độ({self.stats_success}/{self.stats_total}):")

    def update_table_row(self, task_id, col, text):
        for i in range(self.table_tasks.rowCount()):
            item = self.table_tasks.item(i, 1)
            if item and item.data(Qt.ItemDataRole.UserRole) == task_id:
                if col == 4:
                    self.table_tasks.removeCellWidget(i, col)
                    result_item = QTableWidgetItem(text)
                    result_item.setFlags(result_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.table_tasks.setItem(i, col, result_item)
                    result_widget = self.create_result_widget(text)
                    if result_widget:
                        self.table_tasks.setCellWidget(i, col, result_widget)
                        self.table_tasks.setRowHeight(i, 64)
                elif col == 5:
                    self.table_tasks.removeCellWidget(i, col)
                    status_item = QTableWidgetItem(text)
                    status_item.setFlags(status_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                    self.table_tasks.setItem(i, col, status_item)
                    result_item = self.table_tasks.item(i, 4)
                    result_path = result_item.text() if result_item else ""
                    if text == "COMPLETED":
                        self.table_tasks.setCellWidget(i, col, self.create_status_widget(text, result_path))
                else:
                    self.table_tasks.removeCellWidget(i, col)
                    self.table_tasks.setItem(i, col, QTableWidgetItem(text))
                self.filter_table()
                break

    def toggle_select_all(self, state=None):
        is_checked = self.chk_select_all.isChecked()
        for i in range(self.table_tasks.rowCount()):
            if not self.table_tasks.isRowHidden(i):
                widget = self.table_tasks.cellWidget(i, 0)
                if widget:
                    chk = widget.findChild(QCheckBox)
                    if chk:
                        chk.setChecked(is_checked)

    def delete_selected_tasks(self):
        db = SessionLocal()
        rows_to_delete = []
        for i in range(self.table_tasks.rowCount() - 1, -1, -1):
            widget = self.table_tasks.cellWidget(i, 0)
            if widget:
                chk = widget.findChild(QCheckBox)
                if chk and chk.isChecked():
                    task_id = self.table_tasks.item(i, 1).data(Qt.ItemDataRole.UserRole)
                    rows_to_delete.append((i, task_id))
                    
        logging.info(f"[UI] Xóa {len(rows_to_delete)} tasks được tích chọn.")
        for row, task_id in rows_to_delete:
            if task_id:
                task = db.query(Task).filter(Task.id == task_id).first()
                if task:
                    logging.info(f"[UI] Xóa task ID {task_id} (Prompt: '{task.prompt}')")
                    db.delete(task)
                if task_id in self.current_session_task_ids:
                    self.current_session_task_ids.remove(task_id)
            self.table_tasks.removeRow(row)
            
        db.commit()
        db.close()
        
        # Cập nhật ngược lại TextBox bỏ đi các câu prompts vừa bị xóa
        db = SessionLocal()
        remaining_prompts = []
        for t_id in self.current_session_task_ids:
            t = db.query(Task).filter(Task.id == t_id).first()
            if t:
                remaining_prompts.append(t.prompt)
        db.close()

        self.load_tasks()

    def delete_all_tasks(self):
        logging.info("[UI] Xóa sạch toàn bộ tasks.")
        db = SessionLocal()
        self.table_tasks.setRowCount(0)
        db.query(Task).filter(Task.task_type == "video").delete(synchronize_session=False)
        db.commit()
        db.close()
        
        # Đồng bộ xóa cả TextBox
        self.current_session_task_ids.clear()
        self.current_page = 1
        self.load_tasks()
    def get_config_path(self):
        import os
        src_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        return os.path.join(src_dir, "data", "config_video.json")

    def save_config(self):
        import json
        config = {
            "model": self.combo_model.currentText(),
            "quality_1k": self.chk_720p.isChecked(),
            "quality_2k": self.chk_1080p.isChecked(),
            "quality_4k": self.chk_4k.isChecked(),
            "aspect_ratio": self.combo_ratio.currentText(),
            "images_per_prompt": self.spin_images_per_prompt.value(),
            "threads": self.spin_threads.value(),
            "delay_min": self.spin_delay_min.value(),
            "delay_max": self.spin_delay_max.value(),
            "ref_mode": self.combo_ref_mode.currentText(),
            "lock_seed": self.chk_lock_seed.isChecked(),
            "seed": self.line_seed.text(),
        }
        try:
            with open(self.get_config_path(), "w", encoding="utf-8") as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            logging.info("[Config] Lưu cấu hình thành công.")
        except Exception as e:
            logging.error(f"[Config] Lỗi khi lưu cấu hình: {str(e)}")

    def load_config(self):
        import json
        config_file = self.get_config_path()
        if not os.path.exists(config_file):
            return
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                config = json.load(f)
            
            self.combo_model.blockSignals(True)
            self.chk_720p.blockSignals(True)
            self.chk_1080p.blockSignals(True)
            self.chk_4k.blockSignals(True)
            self.combo_ratio.blockSignals(True)
            self.spin_images_per_prompt.blockSignals(True)
            self.spin_threads.blockSignals(True)
            self.spin_delay_min.blockSignals(True)
            self.spin_delay_max.blockSignals(True)
            self.combo_ref_mode.blockSignals(True)
            self.chk_lock_seed.blockSignals(True)
            self.line_seed.blockSignals(True)
            
            self.combo_model.setCurrentText(config.get("model", "Nano Banana Pro"))
            self.chk_720p.setChecked(config.get("quality_1k", True))
            self.chk_1080p.setChecked(config.get("quality_2k", False))
            self.chk_4k.setChecked(config.get("quality_4k", False))
            self.combo_ratio.setCurrentText(config.get("aspect_ratio", "16:9 Ngang"))
            self.spin_images_per_prompt.setValue(config.get("images_per_prompt", 1))
            self.spin_threads.setValue(config.get("threads", 1))
            self.spin_delay_min.setValue(config.get("delay_min", 10))
            self.spin_delay_max.setValue(config.get("delay_max", 20))
            self.combo_ref_mode.setCurrentText(config.get("ref_mode", "Mặc định"))
            self.chk_lock_seed.setChecked(config.get("lock_seed", False))
            self.line_seed.setText(config.get("seed", ""))
            
            self.combo_model.blockSignals(False)
            self.chk_720p.blockSignals(False)
            self.chk_1080p.blockSignals(False)
            self.chk_4k.blockSignals(False)
            self.combo_ratio.blockSignals(False)
            self.spin_images_per_prompt.blockSignals(False)
            self.spin_threads.blockSignals(False)
            self.spin_delay_min.blockSignals(False)
            self.spin_delay_max.blockSignals(False)
            self.combo_ref_mode.blockSignals(False)
            self.chk_lock_seed.blockSignals(False)
            self.line_seed.blockSignals(False)
            
            logging.info("[Config] Tải cấu hình cũ thành công.")
        except Exception as e:
            logging.error(f"[Config] Lỗi khi tải cấu hình: {str(e)}")

    def run_selected_tasks(self):
        logging.info("[UI] Nhấn Chạy mục chọn.")
        
        db = SessionLocal()
        active_account = db.query(Account).filter(Account.is_active == True).order_by(Account.position.asc()).first()
        db.close()
        if not active_account:
            QMessageBox.warning(self, "Thiếu tài khoản", "Vui lòng thêm hoặc kích hoạt ít nhất một tài khoản ở tab 'Cài đặt hệ thống'!")
            return

        selected_task_ids = []
        db = SessionLocal()
        for i in range(self.table_tasks.rowCount()):
            widget = self.table_tasks.cellWidget(i, 0)
            if widget:
                chk = widget.findChild(QCheckBox)
                if chk and chk.isChecked():
                    task_id = self.table_tasks.item(i, 1).data(Qt.ItemDataRole.UserRole)
                    if task_id:
                        selected_task_ids.append(task_id)
                        task = db.query(Task).filter(Task.id == task_id).first()
                        if task:
                            task.status = "PENDING"
                            task.result_path = None
                            if getattr(self, '_current_selected_session_id', None) is None:
                                self._current_selected_session_id = task.session_id
        db.commit()
        db.close()
        
        if not selected_task_ids:
            QMessageBox.warning(self, "Thiếu dữ liệu", "Vui lòng tích chọn ít nhất một công việc trong bảng để chạy!")
            return
            
        self.load_tasks()
        
        import os
        save_path = ""
        sess_id = getattr(self, '_current_selected_session_id', None)
        if sess_id:
            db = SessionLocal()
            sess = db.query(VideoSession).filter(VideoSession.id == sess_id).first()
            if sess:
                save_path = sess.save_dir.strip() if sess.save_dir else ""
                if not save_path:
                    src_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
                    save_path = os.path.join(src_dir, "output", sess.name.replace(" ", "_"))
            db.close()
        self._current_selected_session_id = None
        
        if not save_path:
            QMessageBox.warning(self, "Lỗi", "Không xác định được thư mục lưu!")
            return

        # Reset stats for this run
        self.stats_total = len(selected_task_ids)
        self.stats_processed = 0
        self.stats_success = 0
        self.stats_failure = 0
        self.update_stats_display()
        
        self.base_config = {
            "model": self.combo_model.currentText(),
            "quality": [q.text() for q in [self.chk_720p, self.chk_1080p, self.chk_4k] if q.isChecked()],
            "aspect_ratio": self.combo_ratio.currentText(),
            "images_per_prompt": self.spin_images_per_prompt.value(),
            "save_path": save_path
        }
        
        for i in range(self.table_tasks.rowCount()):
            task_id = self.table_tasks.item(i, 1).data(Qt.ItemDataRole.UserRole)
            if task_id in selected_task_ids:
                stt = self.table_tasks.item(i, 1).text()
                final_name = stt
                
                if not any(t['task_id'] == task_id for t in self.task_queue):
                    if not any(w.task_id == task_id and w.isRunning() for w in self.workers):
                        self.task_queue.append({
                            'task_id': task_id,
                            'final_name': final_name,
                            'target': "labs.google/video"
                        })
                        
        if not self.queue_timer.isActive():
            self.set_running_buttons_enabled()
            self.queue_timer.start(1000)
            self.process_queue()

    def rerun_failed_tasks(self):
        logging.info("[UI] Nhấn Chạy lại lỗi.")
        
        db = SessionLocal()
        active_account = db.query(Account).filter(Account.is_active == True).order_by(Account.position.asc()).first()
        db.close()
        if not active_account:
            QMessageBox.warning(self, "Thiếu tài khoản", "Vui lòng thêm hoặc kích hoạt ít nhất một tài khoản ở tab 'Cài đặt hệ thống'!")
            return

        failed_task_ids = []
        db = SessionLocal()
        for i in range(self.table_tasks.rowCount()):
            status_item = self.table_tasks.item(i, 5)
            if status_item and "ERROR" in status_item.text().upper():
                task_id = self.table_tasks.item(i, 1).data(Qt.ItemDataRole.UserRole)
                if task_id:
                    failed_task_ids.append(task_id)
                    task = db.query(Task).filter(Task.id == task_id).first()
                    if task:
                        task.status = "PENDING"
                        task.result_path = None
                        if getattr(self, '_current_rerun_session_id', None) is None:
                            self._current_rerun_session_id = task.session_id
        db.commit()
        db.close()
        
        if not failed_task_ids:
            QMessageBox.warning(self, "Thông báo", "Không có công việc nào bị lỗi để chạy lại!")
            return
            
        for i in range(self.table_tasks.rowCount()):
            task_id = self.table_tasks.item(i, 1).data(Qt.ItemDataRole.UserRole)
            if task_id in failed_task_ids:
                self.update_table_row(task_id, 5, "PENDING")
                
        QMessageBox.information(self, "Thành công", f"Đã chuyển {len(failed_task_ids)} công việc lỗi sang trạng thái PENDING.")

    def load_real_data(self):
        logging.info("[UI] Nhấn Load dữ liệu thực tế (Chạy theo phiên).")
        import os
        db = SessionLocal()
        
        updated_count = 0
        sessions = db.query(VideoSession).all()
        
        src_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        
        for sess in sessions:
            save_path = sess.save_dir.strip() if sess.save_dir else ""
            if not save_path:
                save_path = os.path.join(src_dir, "output", sess.name.replace(" ", "_"))
            
            if not os.path.exists(save_path):
                continue
                
            tasks = db.query(Task).filter(Task.session_id == sess.id, Task.task_type == "video").order_by(Task.id.asc()).all()
            for idx, task in enumerate(tasks):
                stt = str(idx + 1)
                
                image_exists = False
                found_path = None
                
                possible_names = [
                    f"{stt}.png", f"{stt}.jpg", 
                    f"{stt}_2K.png", f"{stt}_2K.jpg", 
                    f"{stt}_4K.png", f"{stt}_4K.jpg", 
                    f"{stt}_1K.png", f"{stt}_1K.jpg"
                ]
                
                for name in possible_names:
                    p = os.path.join(save_path, name)
                    if os.path.exists(p):
                        image_exists = True
                        found_path = p
                        break
                        
                if image_exists:
                    if task.status != "COMPLETED" or task.result_path != found_path:
                        task.status = "COMPLETED"
                        task.result_path = found_path
                        updated_count += 1
                else:
                    if task.status == "COMPLETED":
                        task.status = "PENDING"
                        task.result_path = None
                        updated_count += 1
                        
        db.commit()
        db.close()
        
        self.load_sessions()
        self.load_tasks()
        QMessageBox.information(self, "Thành công", f"Đã quét và cập nhật trạng thái {updated_count} ảnh từ tất cả các phiên!")
